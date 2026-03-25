from __future__ import annotations

import re
from pathlib import Path
from zipfile import BadZipFile

from lxml import etree
from markdown_it import MarkdownIt
from openpyxl import load_workbook
from pypdf import PdfReader

from spec2cov.config import FilterConfig

PARAGRAPH_SPLIT_RE = re.compile(r"\n\s*\n", re.MULTILINE)
WORD_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_\-]*")
PDF_REFERENCE_RE = re.compile(r"(?i)(?:\[[^\]]*\]\([^)]*\.pdf[^)]*\)|\b[^\s)\]\"']+\.pdf\b)")


def normalize_text(text: str) -> str:
    return "\n".join(line.rstrip() for line in text.replace("\r\n", "\n").splitlines()).strip()


def normalize_match_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def extract_terms(text: str) -> set[str]:
    return {normalize_match_key(token) for token in WORD_RE.findall(text) if normalize_match_key(token)}


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    normalized_headers = [header.strip() or "column" for header in headers]
    lines = [
        "| " + " | ".join(normalized_headers) + " |",
        "| " + " | ".join("---" for _ in normalized_headers) + " |",
    ]
    for row in rows:
        padded = list(row[: len(normalized_headers)])
        while len(padded) < len(normalized_headers):
            padded.append("")
        lines.append("| " + " | ".join(cell.replace("\n", " ").strip() for cell in padded) + " |")
    return "\n".join(lines)


def extract_hvp_text(path: Path) -> list[dict]:
    text = normalize_text(path.read_text(encoding="utf-8", errors="ignore"))
    return [{"type": "hvp", "name": path.stem, "content": text, "metadata": {"source_type": path.suffix.lower()}}] if text else []


def extract_ralf_text(path: Path) -> list[dict]:
    text = normalize_text(path.read_text(encoding="utf-8", errors="ignore"))
    return [{"type": "cover", "name": path.stem, "content": text, "metadata": {"source_type": "ralf", "normalized_keywords": sorted(extract_terms(text))[:200]}}] if text else []


def extract_xml_plan(path: Path) -> list[dict]:
    parser = etree.XMLParser(recover=True)
    tree = etree.parse(str(path), parser)
    rows: list[list[str]] = []
    for node in tree.iter():
        if not isinstance(node.tag, str):
            continue
        value = normalize_text(" ".join(part.strip() for part in node.itertext() if part and part.strip()))
        if not value:
            continue
        rows.append([tree.getpath(node), value])
    if not rows:
        return []
    text = markdown_table(["path", "value"], rows)
    return [{"type": "plan", "name": path.stem, "content": text, "metadata": {"source_type": "xml", "content_format": "markdown_table"}}]


def extract_xlsx_plan(path: Path) -> list[dict]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except BadZipFile:
        return []

    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        normalized_rows = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in rows
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        if not normalized_rows:
            continue
        header = normalized_rows[0]
        body = normalized_rows[1:] or [["" for _ in header]]
        sections.append(f"# Sheet: {sheet.title}")
        sections.append(markdown_table(header, body))
    text = normalize_text("\n\n".join(sections))
    return [{"type": "plan", "name": path.stem, "content": text, "metadata": {"source_type": "xlsx", "content_format": "markdown_table"}}] if text else []


def extract_pdf_spec(path: Path, keyword_terms: set[str] | None = None, min_chars: int = 0) -> list[dict]:
    reader = PdfReader(str(path))
    sections: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = normalize_text(page.extract_text() or "")
        if not text:
            continue
        sections.append(f"## Page {index}\n\n{text}")
    content = normalize_text("\n\n".join(sections))
    if not content:
        return []
    if not keyword_terms:
        return [
            {
                "type": "spec",
                "name": path.stem,
                "content": content,
                "metadata": {
                    "source_type": "pdf",
                    "content_format": "markdown",
                    "extractor": "pypdf",
                    "page_count": len(reader.pages),
                },
            }
        ]
    return extract_textual_spec_sections(content, path.stem, keyword_terms, min_chars, source_type="pdf")


def build_spec_seed_terms(existing_artifacts: list[dict], filter_config: FilterConfig) -> set[str]:
    seeds = {normalize_match_key(keyword) for keyword in filter_config.md_spec_keywords}
    for artifact in existing_artifacts:
        if artifact["type"] in {"plan", "hvp", "cover"}:
            seeds.update(extract_terms(artifact["content"]))
    return {seed for seed in seeds if seed}


def _build_cover_keyword_terms(
    existing_artifacts: list[dict],
    *,
    metadata_fields: tuple[str, ...],
    include_content_terms: bool,
) -> set[str]:
    keywords: set[str] = set()
    for artifact in existing_artifacts:
        if artifact["type"] != "cover":
            continue
        metadata = artifact.get("metadata", {})
        for field in metadata_fields:
            for value in metadata.get(field, []):
                normalized = normalize_match_key(value)
                if normalized:
                    keywords.add(normalized)
        if include_content_terms:
            keywords.update(extract_terms(artifact["content"]))
    return keywords


def build_spec_keyword_terms(existing_artifacts: list[dict]) -> set[str]:
    return _build_cover_keyword_terms(
        existing_artifacts,
        metadata_fields=(
            "normalized_keywords",
            "covergroup_names",
            "coverpoint_names",
            "coverpoint_targets",
            "coverpoint_signal_terms",
            "cross_terms",
        ),
        include_content_terms=True,
    )


def build_dut_keyword_terms(existing_artifacts: list[dict]) -> set[str]:
    return _build_cover_keyword_terms(
        existing_artifacts,
        metadata_fields=("coverpoint_signal_terms",),
        include_content_terms=False,
    )


def markdown_mentions_pdf(path: Path) -> bool:
    source = path.read_text(encoding="utf-8", errors="ignore")
    return PDF_REFERENCE_RE.search(source) is not None


def _paragraph_windows(paragraphs: list[str], index: int) -> str:
    start = max(index - 1, 0)
    end = min(index + 2, len(paragraphs))
    return normalize_text("\n\n".join(paragraphs[start:end]))


def extract_textual_spec_sections(content: str, name_prefix: str, keyword_terms: set[str], min_chars: int, source_type: str) -> list[dict]:
    paragraphs = [normalize_text(block) for block in PARAGRAPH_SPLIT_RE.split(content) if normalize_text(block)]
    if not keyword_terms:
        return []

    results: list[dict] = []
    emitted: set[str] = set()
    for index, paragraph in enumerate(paragraphs, start=1):
        terms = extract_terms(paragraph)
        overlap = sorted(terms & keyword_terms)
        if not overlap:
            continue
        window = _paragraph_windows(paragraphs, index - 1)
        if len(window) < min_chars or window in emitted:
            continue
        emitted.add(window)
        results.append(
            {
                "type": "spec",
                "name": f"{name_prefix}-{index}",
                "content": window,
                "metadata": {"matched_terms": overlap[:20], "source_type": source_type},
            }
        )
    return results


def extract_markdown_spec(path: Path, seed_terms: set[str], min_chars: int) -> list[dict]:
    source = path.read_text(encoding="utf-8", errors="ignore")
    markdown = MarkdownIt()
    rendered = markdown.render(source)
    rendered_text = re.sub(r"<[^>]+>", " ", rendered)
    return extract_textual_spec_sections(rendered_text, path.stem, seed_terms, min_chars, source_type="md")
